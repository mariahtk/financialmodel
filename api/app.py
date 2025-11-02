# api/process_excel.py
import os
import shutil
from fastapi import FastAPI, UploadFile, File
from fastapi.responses import FileResponse
from openpyxl import load_workbook

app = FastAPI()

# Base model path (make sure this exists in your project)
BASE_MODEL = "Bespoke Model - US - v2.xlsm"

# Map of input cells → model cells
CELL_MAP = {
    "F7": "E6",   # Address
    "F13": "E12", # Latitude
    "F15": "E14", # Longitude
    "F29": "E34", # Square footage
    "F37": "K10", # Market rent
    "F54": "K34", # Yes/No value
    "F56": "K36"  # Number of floors
}

@app.post("/process_excel/")
async def process_excel(file: UploadFile = File(...)):
    try:
        # 1. Save uploaded file
        input_path = f"/tmp/{file.filename}"
        with open(input_path, "wb") as f:
            f.write(await file.read())

        # 2. Load workbooks
        try:
            wb_input = load_workbook(input_path, keep_vba=True)
        except Exception as e:
            return {"error": f"Failed to open uploaded input file: {e}"}

        try:
            wb_model = load_workbook(BASE_MODEL, keep_vba=True)
        except Exception as e:
            return {"error": f"Failed to open base model file: {e}"}

        # 3. Access sheets
        try:
            ws_input = wb_input["Sales Team Input Sheet"]
        except KeyError:
            return {"error": "Input sheet 'Sales Team Input Sheet' not found in uploaded file."}

        try:
            ws_model = wb_model["Sales Team Input Sheet"]
        except KeyError:
            return {"error": "Sheet 'Sales Team Input Sheet' not found in base model."}

        # 4. Copy mapped cell values
        for input_cell, model_cell in CELL_MAP.items():
            try:
                value = ws_input[input_cell].value
                if value is not None:
                    ws_model[model_cell].value = value
            except Exception as e:
                return {"error": f"Failed to copy {input_cell} to {model_cell}: {e}"}

        # 5. Handle market rent dropdown logic (optional)
        try:
            market_rent = ws_input["F37"].value
            if market_rent is not None:
                if float(market_rent) == 15:
                    ws_model["K10"].value = "15 - 20"
                elif float(market_rent) == 20:
                    ws_model["K10"].value = "20 - 25"
        except Exception:
            pass

        # 6. Save output
        output_path = f"/tmp/Processed_Model.xlsm"
        try:
            wb_model.save(output_path)
        except Exception as e:
            return {"error": f"Failed to save processed model: {e}"}

        # 7. Return as downloadable file
        return FileResponse(
            path=output_path,
            filename="Processed_Model.xlsm",
            media_type="application/vnd.ms-excel.sheet.macroEnabled.12"
        )

    except Exception as e:
        return {"error": f"Unexpected error: {e}"}
