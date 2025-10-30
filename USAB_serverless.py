import openpyxl

def run_model(input_path, model_path, output_path):
    """
    Reads the uploaded input sheet and copies values to the model.
    Safe for serverless environment using /tmp/.
    """
    # Load input sheet
    wb_input = openpyxl.load_workbook(input_path, data_only=True)
    ws_input = wb_input.active  # or use ws_input = wb_input["Sheet1"]

    # Load model template (keep macros)
    wb_model = openpyxl.load_workbook(model_path, keep_vba=True)
    ws_model = wb_model.active  # or use ws_model = wb_model["Sheet1"]

    # Example: copy column A values from input to model
    for row in range(2, ws_input.max_row + 1):
        val = ws_input[f"A{row}"].value
        if val is not None:
            ws_model[f"A{row}"].value = val

    # Save output
    wb_model.save(output_path)
